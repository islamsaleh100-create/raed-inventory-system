from __future__ import annotations

import os
import sys
from dataclasses import dataclass

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import load_workbook

from app.core.security import get_password_hash
from app.database import SessionLocal
from app.models import (
    AreaManagerAssignment,
    Branch,
    Brand,
    KitchenSection,
    KitchenSectionAssignment,
    Role,
    RoleName,
    User,
    UserRole,
    UserStatus,
    Warehouse,
)


# Local fallback path; staging/CI must set PERMISSION_MATRIX_WORKBOOK.
_DEFAULT_MATRIX_WORKBOOK = r"C:\Users\islam\Downloads\raed_user_matrix_permissions.xlsx"


def matrix_workbook_path() -> str:
    p = (os.environ.get("PERMISSION_MATRIX_WORKBOOK") or "").strip()
    return p if p else _DEFAULT_MATRIX_WORKBOOK


def matrix_seed_password() -> str:
    v = (os.environ.get("PERMISSION_MATRIX_PASSWORD") or "").strip()
    return v if v else "Raed@2025"


ROLE_MAP: dict[str, tuple[RoleName, ...]] = {
    "Super Admin": (RoleName.super_admin,),
    "Admin": (RoleName.admin,),
    "Area Manager": (RoleName.area_manager,),
    # Keep branch accounts operational for current app behavior.
    "Branch User": (RoleName.branch_user, RoleName.branch_manager),
    "Kitchen Section Manager": (RoleName.kitchen_section_manager,),
    "Kitchen Manager": (RoleName.kitchen_manager,),
    "Warehouse Manager": (RoleName.warehouse_manager,),
    "Warehouse User": (RoleName.warehouse_user,),
    "Delivery User": (RoleName.delivery_user,),
}


ROLE_META: dict[RoleName, tuple[str, str | None]] = {
    RoleName.super_admin: ("Super Administrator", "Full system access"),
    RoleName.admin: ("System Administrator", "Administrative access"),
    RoleName.branch_user: ("Branch User", "Branch request entry"),
    RoleName.branch_manager: ("Branch Manager", "Branch operations and employee management"),
    RoleName.area_manager: ("Area Manager", "City + brand scoped approval"),
    RoleName.kitchen_manager: ("Kitchen Manager", "Legacy kitchen overview role"),
    RoleName.kitchen_section_manager: ("Kitchen Section Manager", "Kitchen execution scoped by section"),
    RoleName.warehouse_manager: ("Warehouse Manager", "Warehouse oversight"),
    RoleName.warehouse_user: ("Warehouse User", "Warehouse execution"),
    RoleName.delivery_user: ("Delivery User", "Delivery execution"),
}


BRANCH_NAME_TO_CODE: dict[str, str] = {
    "onda 1 - arkan": "BR-DM-ON-ARKAN",
    "onda 13 - al malqa": "BR-RY-ON-MALQA",
    "onda 14 - hassa": "BR-DM-ON-HASSA",
    "onda 16 - najmah": "BR-DM-ON-NAJMA",
    "onda 18 - al midra gym": "BR-DM-ON-MIDRA",
    "onda 2 - hoqail": "BR-DM-ON-HOQAI",
    "onda 4 - sefarat": "BR-RY-ON-SEFAR",
    "onda 5 - muowasat": "BR-DM-ON-MUOWA",
    "onda 9 - ras tanura": "BR-DM-ON-RASTN",
    "onda dau university": "BR-DM-ON-DAU",
    "pizza 1 - al khobar": "BR-DM-RN-KHOBR",
    "pizza 10 - mazaar": "BR-DM-RN-MAZAR",
    "pizza 15 - ras tanura": "BR-DM-RN-RASTN",
    "pizza 3 - arkan": "BR-DM-RN-ARKAN",
    "pizza 4 - riyadh takhasosy": "BR-RY-RN-TAKHS",
    "pizza 5 - al ulaya": "BR-RY-RN-ULAYA",
    "pizza 6 - riyadh nada": "BR-RY-RN-NADA",
    "pizza 7 - aramco": "BR-DM-RN-ARAMC",
    "pizza 9 - al azizia": "BR-DM-RN-AZIZI",
    "ronaldos dau university": "BR-DM-RN-DAU",
    "shawerma 1 - khobar": "BR-DM-SH-KHOBR",
    "shawarma 1 - khobar": "BR-DM-SH-KHOBR",
    "shawerma 4 - arkan": "BR-DM-SH-ARKAN",
    "shawarma 4 - arkan": "BR-DM-SH-ARKAN",
    "shawerma - olaya": "BR-RY-SH-OLAYA",
    "shawarma - olaya": "BR-RY-SH-OLAYA",
}


def normalize(value: str | None) -> str:
    return (value or "").strip().lower()


def get_or_create_role(db, role_name: RoleName) -> Role:
    row = db.query(Role).filter(Role.name == role_name).first()
    if row:
        return row
    display_name, description = ROLE_META.get(role_name, (role_name.value.replace("_", " ").title(), None))
    row = Role(name=role_name, display_name=display_name, description=description)
    db.add(row)
    db.flush()
    return row


def ensure_user_roles(db, user: User, roles: tuple[RoleName, ...]) -> None:
    existing_role_ids = {ur.role_id for ur in user.user_roles}
    for role_name in roles:
        role = get_or_create_role(db, role_name)
        if role.id not in existing_role_ids:
            db.add(UserRole(user_id=user.id, role_id=role.id))
    db.flush()


def get_or_create_user(db, *, username: str, full_name: str, email: str, status: UserStatus, roles: tuple[RoleName, ...]) -> User:
    user = db.query(User).filter(User.username == username).first()
    if not user:
        user = User(
            username=username,
            email=email,
            full_name=full_name,
            hashed_password=get_password_hash(matrix_seed_password()),
            status=status,
            is_deleted=False,
        )
        db.add(user)
        db.flush()
    user.email = email
    user.full_name = full_name
            user.hashed_password = get_password_hash(matrix_seed_password())
    user.status = status
    user.is_deleted = False
    ensure_user_roles(db, user, roles)
    return user


def get_branch_by_name(db, entity_name: str) -> Branch | None:
    code = BRANCH_NAME_TO_CODE.get(normalize(entity_name))
    if not code:
        return None
    return db.query(Branch).filter(Branch.branch_code == code, Branch.is_deleted == False).first()


def get_brand(db, name: str) -> Brand | None:
    return db.query(Brand).filter(Brand.name == name).first()


def get_warehouse(db, city: str) -> Warehouse | None:
    normalized = normalize(city)
    if normalized == "riyadh":
        row = db.query(Warehouse).filter(Warehouse.warehouse_code == "WH-RY-1", Warehouse.is_deleted == False).first()
        if row:
            return row
    if normalized == "dammam":
        row = db.query(Warehouse).filter(Warehouse.warehouse_code == "WH-DM-1", Warehouse.is_deleted == False).first()
        if row:
            return row
    return db.query(Warehouse).filter(Warehouse.is_deleted == False, Warehouse.active == True).order_by(Warehouse.id.asc()).first()


def get_section(section_label: str) -> str | None:
    lower = normalize(section_label)
    if "meat & chicken" in lower:
        return "Meat & Chicken"
    if "bakery & sweets" in lower:
        return "Bakery & Sweets"
    if "pizza" in lower:
        return "Pizza"
    return None


def ensure_area_assignment(db, user: User, city: str, brand_name: str) -> None:
    brand = get_brand(db, brand_name)
    if not brand:
        return
    row = db.query(AreaManagerAssignment).filter(
        AreaManagerAssignment.user_id == user.id,
        AreaManagerAssignment.city == city,
        AreaManagerAssignment.brand_id == brand.id,
        AreaManagerAssignment.active == True,
    ).first()
    if row:
        return
    db.add(AreaManagerAssignment(user_id=user.id, city=city, brand_id=brand.id, active=True))
    db.flush()


def ensure_section_assignment(db, user: User, section_name: str, service_city: str | None = None) -> None:
    section = db.query(KitchenSection).filter(KitchenSection.name == section_name).first()
    if not section:
        return
    row = db.query(KitchenSectionAssignment).filter(
        KitchenSectionAssignment.user_id == user.id,
        KitchenSectionAssignment.kitchen_section_id == section.id,
        KitchenSectionAssignment.active == True,
    ).first()
    city_val = (service_city or "").strip() or None
    if row:
        if city_val and row.service_city != city_val:
            row.service_city = city_val
            db.flush()
        return
    db.add(
        KitchenSectionAssignment(
            user_id=user.id,
            kitchen_section_id=section.id,
            active=True,
            service_city=city_val,
        )
    )
    db.flush()


@dataclass(frozen=True)
class MatrixRow:
    username: str
    display_name: str
    role_label: str
    city: str
    brand_scope: str
    entity_type: str
    assignment: str
    status: str


def iter_matrix_rows(workbook_path: str) -> list[MatrixRow]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    ws = wb["Users"]
    rows: list[MatrixRow] = []
    for record in ws.iter_rows(min_row=2, values_only=True):
        if not record or not record[0]:
            continue
        rows.append(
            MatrixRow(
                username=str(record[0]).strip(),
                display_name=str(record[1] or "").strip(),
                role_label=str(record[2] or "").strip(),
                city=str(record[3] or "").strip(),
                brand_scope=str(record[4] or "").strip(),
                entity_type=str(record[5] or "").strip(),
                assignment=str(record[6] or "").strip(),
                status=str(record[7] or "").strip(),
            )
        )
    return rows


def main() -> int:
    wb_path = matrix_workbook_path()
    print(f"matrix_workbook={wb_path}")
    if not os.path.isfile(wb_path):
        print(
            "ERROR: workbook file not found. Set PERMISSION_MATRIX_WORKBOOK to the absolute path "
            "of raed_user_matrix_permissions.xlsx (see STAGING_HANDOFF_REPORT.md).",
            file=sys.stderr,
        )
        return 1
    db = SessionLocal()
    try:
        rows = iter_matrix_rows(wb_path)
        created = 0
        updated = 0
        warnings: list[str] = []
        for row in rows:
            roles = ROLE_MAP.get(row.role_label)
            if not roles:
                warnings.append(f"unknown role label: {row.username} -> {row.role_label}")
                continue
            status = UserStatus.active if normalize(row.status) == "active" else UserStatus.inactive
            email = f"{row.username}@raed.com"
            existed = db.query(User).filter(User.username == row.username).first() is not None
            user = get_or_create_user(
                db,
                username=row.username,
                full_name=row.display_name or row.username,
                email=email,
                status=status,
                roles=roles,
            )
            user.branch_id = None
            user.warehouse_id = None
            if not existed:
                created += 1
            else:
                updated += 1

            if row.entity_type == "Branch":
                branch = get_branch_by_name(db, row.assignment)
                if branch:
                    user.branch_id = branch.id
                else:
                    warnings.append(f"branch not found for {row.username}: {row.assignment}")
            elif row.entity_type == "Warehouse":
                wh = get_warehouse(db, row.city)
                if wh:
                    user.warehouse_id = wh.id
                else:
                    warnings.append(f"warehouse not found for {row.username}: {row.city}")
            elif row.entity_type == "Delivery":
                wh = get_warehouse(db, row.city)
                if wh:
                    user.warehouse_id = wh.id
                else:
                    warnings.append(f"delivery warehouse not found for {row.username}: {row.city}")
            elif row.entity_type == "Area":
                brands = ["Onda", "Ronaldos", "Shawarma", "Griddle"] if normalize(row.brand_scope) == "all" else [x.strip() for x in row.brand_scope.split(",") if x.strip()]
                for brand_name in brands:
                    ensure_area_assignment(db, user, row.city, brand_name)
            elif row.entity_type == "Kitchen Section":
                section_name = get_section(row.assignment)
                if section_name:
                    ensure_section_assignment(db, user, section_name, service_city=row.city or None)
                else:
                    warnings.append(f"kitchen section not resolved for {row.username}: {row.assignment}")

        db.commit()
        print(f"matrix_users_total={len(rows)}")
        print(f"users_created={created}")
        print(f"users_updated={updated}")
        print(f"warnings_count={len(warnings)}")
        for line in warnings:
            print(f"WARNING: {line}")
        return 0
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


if __name__ == "__main__":
    raise SystemExit(main())
