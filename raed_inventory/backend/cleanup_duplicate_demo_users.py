from __future__ import annotations

import argparse
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from openpyxl import load_workbook

from app.database import SessionLocal
from app.models import User, UserStatus


WORKBOOK_PATH = r"C:\Users\islam\Downloads\raed_user_matrix_permissions.xlsx"

# Keep official matrix users plus a very small set of canonical operational/demo accounts
# that still help current runtime operation.
EXTRA_KEEP_USERNAMES: set[str] = {
    "branch_griddle",
    "kitchen_manager",
    "meat_manager",
    "bakery_sweets_manager",
    "pizza_manager",
    "warehouse_user",
    "delivery_user",
}


def load_matrix_usernames() -> set[str]:
    wb = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
    ws = wb["Users"]
    rows: set[str] = set()
    for record in ws.iter_rows(min_row=2, values_only=True):
        if record and record[0]:
            rows.add(str(record[0]).strip())
    return rows


def main() -> int:
    parser = argparse.ArgumentParser(description="Deactivate duplicate demo users while keeping matrix + essential canonicals.")
    parser.add_argument("--dry-run", action="store_true", help="Print changes without commit.")
    args = parser.parse_args()

    keep_usernames = load_matrix_usernames() | EXTRA_KEEP_USERNAMES
    db = SessionLocal()
    try:
        users = db.query(User).filter(User.is_deleted == False).all()  # noqa: E712
        deactivated: list[str] = []
        kept: list[str] = []
        for user in users:
            if user.username in keep_usernames:
                kept.append(user.username)
                continue
            if user.status != UserStatus.inactive:
                deactivated.append(user.username)
                if not args.dry_run:
                    user.status = UserStatus.inactive
        if args.dry_run:
            db.rollback()
        else:
            db.commit()
        print(f"keep_count={len(kept)}")
        print(f"deactivated_count={len(deactivated)}")
        for username in sorted(deactivated):
            print(f"deactivated:{username}")
        if args.dry_run:
            print("(dry-run: no database changes committed)")
        return 0
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


if __name__ == "__main__":
    raise SystemExit(main())
