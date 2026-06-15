"""
Export Router — /api/v1/export
Epic 12: Download data as CSV or XLSX

Supports:
  GET /export/inventory-compliance?date_from=&date_to=&format=csv|xlsx
  GET /export/variance-report?branch_id=&date_from=&date_to=&format=csv|xlsx
  GET /export/order-summary?date_from=&date_to=&format=csv|xlsx
  GET /export/stock/branches/{branch_id}?format=csv|xlsx
  GET /export/stock/warehouses/{warehouse_id}?format=csv|xlsx
  GET /export/ledger/branches/{branch_id}?format=csv|xlsx
"""
import csv
import io
from datetime import date, datetime, timedelta
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from sqlalchemy import and_
from sqlalchemy.orm import Session, joinedload

from app.core.area_manager_scope import get_area_manager_branch_ids
from app.core.auth import (
    can_access_branch,
    can_access_warehouse,
    get_current_active_user,
    get_user_roles,
    is_platform_admin,
    require_roles,
)
from app.database import get_db
from app.models import (
    Branch,
    BranchStock,
    DailyInventory,
    DailyInventoryLine,
    InventoryStatus,
    Item,
    OrderStatus,
    ReplenishmentOrder,
    StockTransaction,
    User,
    Warehouse,
    WarehouseStock,
)
from app.services import ledger_service

router = APIRouter(prefix="/api/v1/export", tags=["Data Export"])

_MGMT = ("branch_manager", "warehouse_manager", "admin", "super_admin")


def _require_branch_export_access(current_user: User, branch_id: int, db: Session) -> None:
    if not can_access_branch(current_user, branch_id, db):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Access denied for this branch export",
        )


def _require_warehouse_export_access(current_user: User, warehouse_id: int) -> None:
    if not can_access_warehouse(current_user, warehouse_id):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="Access denied for this warehouse export",
        )


def _scoped_branch_ids(current_user: User, db: Session, branch_id: Optional[int] = None) -> list[int]:
    roles = get_user_roles(current_user)
    if is_platform_admin(current_user):
        if branch_id is not None:
            return [branch_id]
        return [row.id for row in db.query(Branch.id).filter(Branch.active == True).all()]  # noqa: E712

    if "branch_manager" in roles or "branch_user" in roles:
        if not current_user.branch_id:
            raise HTTPException(status_code=403, detail="Branch user has no branch assignment")
        if branch_id is not None and branch_id != current_user.branch_id:
            raise HTTPException(status_code=403, detail="Access denied for this branch export")
        return [current_user.branch_id]

    if "area_manager" in roles:
        allowed = get_area_manager_branch_ids(current_user, db)
        if branch_id is not None:
            if branch_id not in allowed:
                raise HTTPException(status_code=403, detail="Access denied for this branch export")
            return [branch_id]
        return allowed

    if "warehouse_manager" in roles or "warehouse_user" in roles:
        if branch_id is not None:
            branch = db.query(Branch).filter(Branch.id == branch_id).first()
            if not branch or not can_access_warehouse(current_user, branch.warehouse_id):
                raise HTTPException(status_code=403, detail="Access denied for this branch export")
            return [branch_id]
        if not current_user.warehouse_id:
            raise HTTPException(status_code=403, detail="Warehouse user has no warehouse assignment")
        return [
            row.id
            for row in db.query(Branch.id).filter(
                Branch.active == True,  # noqa: E712
                Branch.warehouse_id == current_user.warehouse_id,
            ).all()
        ]

    raise HTTPException(status_code=403, detail="Export not allowed for this role")


# ──────────────────────────────────────────────────────────────────────────
# helpers
# ──────────────────────────────────────────────────────────────────────────

def _csv_response(rows: list[dict], filename: str) -> StreamingResponse:
    if not rows:
        rows = [{}]
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=rows[0].keys())
    writer.writeheader()
    writer.writerows(rows)
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _xlsx_response(rows: list[dict], filename: str, sheet_name: str = "Sheet1") -> StreamingResponse:
    try:
        import openpyxl
    except ImportError:
        # Fallback to CSV if openpyxl not installed
        return _csv_response(rows, filename.replace(".xlsx", ".csv"))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    if not rows:
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    headers = list(rows[0].keys())
    ws.append(headers)

    # Bold header row
    from openpyxl.styles import Font, PatternFill
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    for row in rows:
        ws.append([row.get(h) for h in headers])

    # Auto-width columns
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _respond(rows: list[dict], filename_base: str, fmt: str, sheet_name: str = "Data") -> StreamingResponse:
    if fmt == "xlsx":
        return _xlsx_response(rows, f"{filename_base}.xlsx", sheet_name)
    return _csv_response(rows, f"{filename_base}.csv")


# ──────────────────────────────────────────────────────────────────────────
# INVENTORY COMPLIANCE EXPORT
# ──────────────────────────────────────────────────────────────────────────

@router.get("/inventory-compliance")
def export_inventory_compliance(
    date_from: date,
    date_to: date,
    branch_id: Optional[int] = None,
    format: str = Query("csv", pattern="^(csv|xlsx)$"),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(*_MGMT)),
):
    delta = (date_to - date_from).days
    if delta < 0 or delta > 90:
        from app.core.errors import AppError
        raise AppError(
            status_code=400,
            error_code="export.invalid_date_range",
            message="Date range must be 0–90 days",
            detail={},
        )

    branch_ids = _scoped_branch_ids(current_user, db, branch_id)
    branches = db.query(Branch).filter(Branch.id.in_(branch_ids if branch_ids else [-1])).all()

    inv_q = db.query(DailyInventory).filter(
        DailyInventory.inventory_date >= date_from,
        DailyInventory.inventory_date <= date_to,
        DailyInventory.branch_id.in_(branch_ids if branch_ids else [-1]),
    )
    inv_map = {(i.branch_id, i.inventory_date): i for i in inv_q.all()}

    dates = [date_from + timedelta(days=d) for d in range(delta + 1)]
    rows = []
    for branch in branches:
        for d in dates:
            inv = inv_map.get((branch.id, d))
            rows.append({
                "branch_id": branch.id,
                "branch_name": branch.branch_name,
                "date": str(d),
                "status": inv.status.value if inv else "missing",
                "inventory_id": inv.id if inv else "",
            })

    return _respond(rows, f"inventory_compliance_{date_from}_{date_to}", format, "Compliance")


# ──────────────────────────────────────────────────────────────────────────
# VARIANCE REPORT EXPORT
# ──────────────────────────────────────────────────────────────────────────

@router.get("/variance-report")
def export_variance_report(
    branch_id: Optional[int] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    critical_only: bool = False,
    format: str = Query("csv", pattern="^(csv|xlsx)$"),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(*_MGMT)),
):
    if branch_id is not None:
        _require_branch_export_access(current_user, branch_id, db)
    elif not is_platform_admin(current_user):
        raise HTTPException(status_code=400, detail="branch_id is required for this export")
    data = ledger_service.get_variance_report(
        db,
        branch_id=branch_id,
        date_from=date_from,
        date_to=date_to,
        critical_only=critical_only,
        page=1,
        page_size=10000,
    )
    rows = data["items"]
    return _respond(rows, "variance_report", format, "Variance")


# ──────────────────────────────────────────────────────────────────────────
# ORDER SUMMARY EXPORT
# ──────────────────────────────────────────────────────────────────────────

@router.get("/order-summary")
def export_order_summary(
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    branch_id: Optional[int] = None,
    warehouse_id: Optional[int] = None,
    format: str = Query("csv", pattern="^(csv|xlsx)$"),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(*_MGMT)),
):
    q = db.query(ReplenishmentOrder)
    roles = get_user_roles(current_user)
    if "branch_manager" in roles or "branch_user" in roles:
        q = q.filter(ReplenishmentOrder.branch_id == current_user.branch_id)
    elif "warehouse_manager" in roles or "warehouse_user" in roles:
        if not current_user.warehouse_id:
            raise HTTPException(status_code=403, detail="Warehouse user has no warehouse assignment")
        q = q.filter(ReplenishmentOrder.warehouse_id == current_user.warehouse_id)
    elif "area_manager" in roles:
        scoped_ids = get_area_manager_branch_ids(current_user, db)
        q = q.filter(ReplenishmentOrder.branch_id.in_(scoped_ids if scoped_ids else [-1]))

    if date_from:
        q = q.filter(ReplenishmentOrder.order_date >= date_from)
    if date_to:
        q = q.filter(ReplenishmentOrder.order_date <= date_to)
    if branch_id:
        _require_branch_export_access(current_user, branch_id, db)
        q = q.filter(ReplenishmentOrder.branch_id == branch_id)
    if warehouse_id:
        _require_warehouse_export_access(current_user, warehouse_id)
        q = q.filter(ReplenishmentOrder.warehouse_id == warehouse_id)

    orders = q.all()
    rows = [
        {
            "order_id": o.id,
            "order_no": o.order_no,
            "branch_id": o.branch_id,
            "warehouse_id": o.warehouse_id,
            "order_type": o.order_type.value if o.order_type else "",
            "status": o.status.value if o.status else "",
            "order_date": str(o.order_date),
            "dispatch_note_no": o.dispatch_note_no or "",
            "created_at": str(o.created_at),
        }
        for o in orders
    ]
    return _respond(rows, "order_summary", format, "Orders")


# ──────────────────────────────────────────────────────────────────────────
# BRANCH STOCK EXPORT
# ──────────────────────────────────────────────────────────────────────────

@router.get("/stock/branches/{branch_id}")
def export_branch_stock(
    branch_id: int,
    format: str = Query("csv", pattern="^(csv|xlsx)$"),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(*_MGMT)),
):
    _require_branch_export_access(current_user, branch_id, db)
    stocks = db.query(BranchStock).options(
        joinedload(BranchStock.item)
    ).filter(BranchStock.branch_id == branch_id).all()

    rows = [
        {
            "item_id": s.item_id,
            "item_code": s.item.item_code if s.item else "",
            "item_name_ar": s.item.item_name_ar if s.item else "",
            "item_name_en": s.item.item_name_en if s.item else "",
            "current_qty": float(s.current_qty),
            "in_transit_qty": float(s.in_transit_qty),
            "reserved_qty": float(s.reserved_qty),
            "reorder_point": float(s.item.reorder_point) if s.item else 0,
            "min_qty": float(s.item.min_qty) if s.item else 0,
        }
        for s in stocks
    ]
    return _respond(rows, f"branch_{branch_id}_stock", format, "Stock")


# ──────────────────────────────────────────────────────────────────────────
# WAREHOUSE STOCK EXPORT
# ──────────────────────────────────────────────────────────────────────────

@router.get("/stock/warehouses/{warehouse_id}")
def export_warehouse_stock(
    warehouse_id: int,
    format: str = Query("csv", pattern="^(csv|xlsx)$"),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(*_MGMT)),
):
    _require_warehouse_export_access(current_user, warehouse_id)
    rows_query = (
        db.query(Item, WarehouseStock)
        .outerjoin(
            WarehouseStock,
            and_(WarehouseStock.item_id == Item.id, WarehouseStock.warehouse_id == warehouse_id),
        )
        .filter(Item.is_deleted == False)
        .order_by(Item.category_id.asc(), Item.item_name_ar.asc(), Item.item_code.asc())
        .all()
    )

    rows = []
    for item, stock in rows_query:
        rows.append({
            "item_id": item.id,
            "item_code": item.item_code,
            "item_name_ar": item.item_name_ar,
            "item_name_en": item.item_name_en,
            "current_qty": float(stock.current_qty) if stock else 0,
            "reserved_qty": float(stock.reserved_qty) if stock else 0,
            "reorder_point": float(item.reorder_point) if item else 0,
        })
    return _respond(rows, f"warehouse_{warehouse_id}_stock", format, "Stock")


# ──────────────────────────────────────────────────────────────────────────
# LEDGER EXPORT
# ──────────────────────────────────────────────────────────────────────────

@router.get("/ledger/branches/{branch_id}")
def export_branch_ledger(
    branch_id: int,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    format: str = Query("csv", pattern="^(csv|xlsx)$"),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_roles(*_MGMT)),
):
    _require_branch_export_access(current_user, branch_id, db)
    data = ledger_service.get_branch_ledger(
        db,
        branch_id=branch_id,
        date_from=date_from,
        date_to=date_to,
        page=1,
        page_size=10000,
    )
    rows = data["items"]
    # Flatten datetime objects
    for row in rows:
        if isinstance(row.get("transaction_date"), datetime):
            row["transaction_date"] = row["transaction_date"].isoformat()
    return _respond(rows, f"branch_{branch_id}_ledger", format, "Ledger")
