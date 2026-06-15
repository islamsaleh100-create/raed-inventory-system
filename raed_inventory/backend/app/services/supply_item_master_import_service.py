from __future__ import annotations

import csv
import hashlib
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl
from sqlalchemy.orm import Session

from app.models import (
    Brand,
    Item,
    ItemBrand,
    ItemCategory,
    ItemType,
    KitchenSection,
    StorageType,
    SupplyDefaultSource,
    SupplySourceType,
    UnitOfMeasure,
)


OFFICIAL_BRANDS = ("Onda", "Ronaldos", "Shawarma", "Griddle")
ALLOWED_BRAND_LABELS = set(OFFICIAL_BRANDS) | {"General", "Shared"}
GENERAL_BRAND_NAMES = set(OFFICIAL_BRANDS)
SHARED_BRAND_NAMES = {"Ronaldos", "Shawarma", "Griddle"}

SOURCE_VALUES = {member.value: member for member in SupplySourceType}
DEFAULT_SOURCE_VALUES = {member.value: member for member in SupplyDefaultSource}
VALID_KITCHEN_SECTIONS = frozenset({"Meat & Chicken", "Bakery & Sweets", "Pizza"})
ITEM_TYPE_VALUES = {
    "RAW": ItemType.raw_material,
    "FINISHED": ItemType.finished_good,
    "BOTH": ItemType.finished_good,
}


@dataclass
class ImportResult:
    rows_read: int
    imported_items: int
    created_items: int
    updated_items: int
    hidden_items: int
    rejected_rows: list[dict[str, Any]]
    invalid_log_path: str

    def as_dict(self) -> dict[str, Any]:
        return {
            "rows_read": self.rows_read,
            "imported_items": self.imported_items,
            "created_items": self.created_items,
            "updated_items": self.updated_items,
            "hidden_items": self.hidden_items,
            "rejected_rows": self.rejected_rows,
            "invalid_log_path": self.invalid_log_path,
        }


def _boolish(value: Any) -> bool:
    if value is None:
        return False
    text = str(value).strip().lower()
    return text in {"1", "true", "yes", "y", "نعم"}


def _text(value: Any) -> str:
    return str(value or "").strip()


def _slugish(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9]+", "-", value).strip("-").upper()
    return cleaned[:12] or "ITEM"


def _category_code(category_name: str) -> str:
    digest = hashlib.sha1(category_name.encode("utf-8")).hexdigest()[:10].upper()
    return f"CAT-{digest}"


def _item_code(brand_label: str, category_name: str, item_name: str) -> str:
    digest = hashlib.sha1(f"{brand_label}|{category_name}|{item_name}".encode("utf-8")).hexdigest()[:10].upper()
    prefix = _slugish(brand_label)[:5]
    return f"SUP-{prefix}-{digest}"


def _final_item_code(brand_label: str, item_name: str) -> str:
    digest = hashlib.sha1(f"FINAL|{brand_label}|{item_name}".encode("utf-8")).hexdigest()[:10].upper()
    prefix = _slugish(brand_label)[:5]
    return f"SUPF-{prefix}-{digest}"


def _parse_item_type(value: Any, *, source_type: SupplySourceType, can_branch_request: bool) -> tuple[ItemType | None, str | None]:
    text = _text(value).upper()
    if not text:
        if source_type == SupplySourceType.KITCHEN:
            return ItemType.finished_good, None
        if can_branch_request:
            return ItemType.finished_good, None
        return ItemType.consumable, None
    if text not in ITEM_TYPE_VALUES:
        return None, f"Invalid item_type '{text}'"
    item_type = ITEM_TYPE_VALUES[text]
    if text == "RAW" and can_branch_request:
        return None, "RAW items cannot be branch-requestable"
    if text == "RAW" and source_type != SupplySourceType.NOT_REQUESTABLE:
        return ItemType.raw_material, None
    return item_type, None


def _validate_source_rules(
    *,
    source_type: SupplySourceType,
    default_source: SupplyDefaultSource | None,
    kitchen_section: str,
    can_branch_request: bool,
) -> str | None:
    if source_type == SupplySourceType.NOT_REQUESTABLE and can_branch_request:
        return "NOT_REQUESTABLE items cannot be branch-requestable"
    if source_type == SupplySourceType.BOTH and default_source is None:
        return "BOTH items require default_source"
    if source_type == SupplySourceType.KITCHEN:
        if not kitchen_section:
            return "Kitchen item requires kitchen_section_id"
        if default_source and default_source != SupplyDefaultSource.KITCHEN:
            return "KITCHEN item cannot default to WAREHOUSE"
    if source_type == SupplySourceType.WAREHOUSE:
        if kitchen_section:
            return "WAREHOUSE item cannot include kitchen section"
        if default_source and default_source != SupplyDefaultSource.WAREHOUSE:
            return "WAREHOUSE item cannot default to KITCHEN"
    if source_type == SupplySourceType.BOTH and default_source == SupplyDefaultSource.KITCHEN and not kitchen_section:
        return "BOTH item with KITCHEN default requires kitchen section"
    return None


def _brand_targets(brand_label: str) -> list[str]:
    if brand_label == "General":
        return list(GENERAL_BRAND_NAMES)
    if brand_label == "Shared":
        return list(SHARED_BRAND_NAMES)
    return [brand_label]


def _sheet_name(workbook: openpyxl.Workbook) -> str:
    if "Classified_Items" in workbook.sheetnames:
        return "Classified_Items"
    if "Items Final" in workbook.sheetnames:
        return "Items Final"
    raise ValueError(f"Unsupported workbook sheets: {', '.join(workbook.sheetnames)}")


def _load_rows(workbook_path: Path) -> tuple[list[dict[str, Any]], list[tuple[int, str]]]:
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    sheet = _sheet_name(wb)
    if sheet == "Items Final":
        return _load_rows_items_final(wb["Items Final"])

    ws = wb["Classified_Items"]
    headers = [cell.value for cell in ws[1]]
    header_index = {name: idx for idx, name in enumerate(headers)}
    required_headers = {
        "Original Row",
        "Brand",
        "POS Category",
        "Item Name",
        "Source Type",
        "Default Source",
        "Kitchen Section",
        "Can Branch Request",
        "Visible in Branch UI",
    }
    missing = sorted(required_headers - set(header_index))
    if missing:
        raise ValueError(f"Workbook is missing required headers: {', '.join(missing)}")

    rows: list[dict[str, Any]] = []
    errors: list[tuple[int, str]] = []
    seen_keys: dict[tuple[str, str], int] = {}

    for excel_row_no, raw in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(value is not None and str(value).strip() != "" for value in raw):
            continue

        row = {name: raw[idx] if idx < len(raw) else None for name, idx in header_index.items()}
        brand_label = _text(row["Brand"])
        category_name = _text(row["POS Category"])
        item_name = _text(row["Item Name"])
        source_text = _text(row["Source Type"]).upper()
        default_text = _text(row["Default Source"]).upper()
        kitchen_section = _text(row["Kitchen Section"])

        duplicate_key = (brand_label.casefold(), item_name.casefold())
        if duplicate_key in seen_keys:
            errors.append((excel_row_no, f"Duplicate item name for brand '{brand_label}'"))
            continue
        seen_keys[duplicate_key] = excel_row_no

        if not brand_label or not category_name or not item_name:
            errors.append((excel_row_no, "Brand, category, and item name are required"))
            continue

        if brand_label not in ALLOWED_BRAND_LABELS:
            errors.append((excel_row_no, f"Unknown brand '{brand_label}'"))
            continue

        if source_text not in SOURCE_VALUES:
            errors.append((excel_row_no, f"Invalid source_type '{source_text}'"))
            continue

        source_type = SOURCE_VALUES[source_text]
        default_source = DEFAULT_SOURCE_VALUES.get(default_text) if default_text else None
        can_branch_request = _boolish(row["Can Branch Request"]) and source_type != SupplySourceType.NOT_REQUESTABLE

        if kitchen_section and kitchen_section not in VALID_KITCHEN_SECTIONS:
            errors.append((excel_row_no, f"Unknown kitchen section '{kitchen_section}'"))
            continue

        rule_error = _validate_source_rules(
            source_type=source_type,
            default_source=default_source,
            kitchen_section=kitchen_section,
            can_branch_request=can_branch_request,
        )
        if rule_error:
            errors.append((excel_row_no, rule_error))
            continue

        item_type, type_error = _parse_item_type(
            row.get("Item Type"),
            source_type=source_type,
            can_branch_request=can_branch_request,
        )
        if type_error:
            errors.append((excel_row_no, type_error))
            continue

        if source_type == SupplySourceType.KITCHEN:
            default_source = SupplyDefaultSource.KITCHEN
        elif source_type == SupplySourceType.WAREHOUSE:
            default_source = SupplyDefaultSource.WAREHOUSE
        elif source_type == SupplySourceType.NOT_REQUESTABLE:
            default_source = SupplyDefaultSource.WAREHOUSE
        elif default_source is None:
            default_source = SupplyDefaultSource.WAREHOUSE

        rows.append(
            {
                "excel_row_no": excel_row_no,
                "brand_label": brand_label,
                "brand_targets": _brand_targets(brand_label),
                "category_name": category_name,
                "item_name": item_name,
                "source_type": source_type,
                "default_source": default_source,
                "kitchen_section_name": kitchen_section or None,
                "can_branch_request": can_branch_request,
                "visible_in_branch_ui": _boolish(row["Visible in Branch UI"]) and source_type != SupplySourceType.NOT_REQUESTABLE,
                "item_type": item_type,
            }
        )

    return rows, errors


def _load_rows_items_final(ws: openpyxl.worksheet.worksheet.Worksheet) -> tuple[list[dict[str, Any]], list[tuple[int, str]]]:
    headers = [cell.value for cell in ws[1]]
    header_index = {name: idx for idx, name in enumerate(headers)}
    required_headers = {
        "Brand",
        "Item Name",
        "Source Type",
        "Default Source",
        "Kitchen Section",
        "Item Type",
    }
    missing = sorted(required_headers - set(header_index))
    if missing:
        raise ValueError(f"Workbook is missing required headers: {', '.join(missing)}")

    rows: list[dict[str, Any]] = []
    errors: list[tuple[int, str]] = []
    seen_keys: dict[tuple[str, str], int] = {}

    for excel_row_no, raw in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(value is not None and str(value).strip() != "" for value in raw):
            continue

        row = {name: raw[idx] if idx < len(raw) else None for name, idx in header_index.items()}
        brand_label = _text(row["Brand"])
        item_name = _text(row["Item Name"])
        source_text = _text(row["Source Type"]).upper()
        default_text = _text(row["Default Source"]).upper()
        kitchen_section = _text(row["Kitchen Section"])
        item_type_text = _text(row["Item Type"]).upper()

        duplicate_key = (brand_label.casefold(), item_name.casefold())
        if duplicate_key in seen_keys:
            errors.append((excel_row_no, f"Duplicate item name for brand '{brand_label}'"))
            continue
        seen_keys[duplicate_key] = excel_row_no

        if brand_label not in OFFICIAL_BRANDS:
            errors.append((excel_row_no, f"Unsupported brand '{brand_label}'"))
            continue
        if not item_name:
            errors.append((excel_row_no, "Item name is required"))
            continue
        if source_text not in SOURCE_VALUES:
            errors.append((excel_row_no, f"Unsupported source_type '{source_text}'"))
            continue

        source_type = SOURCE_VALUES[source_text]
        default_source = DEFAULT_SOURCE_VALUES.get(default_text) if default_text else None
        if default_source is None:
            errors.append((excel_row_no, f"Unsupported default_source '{default_text}'"))
            continue
        if source_type == SupplySourceType.KITCHEN and not kitchen_section:
            errors.append((excel_row_no, "Kitchen item requires kitchen section"))
            continue
        if source_type == SupplySourceType.WAREHOUSE and kitchen_section:
            errors.append((excel_row_no, "Warehouse item cannot include kitchen section"))
            continue
        if source_type == SupplySourceType.WAREHOUSE and default_source != SupplyDefaultSource.WAREHOUSE:
            errors.append((excel_row_no, "WAREHOUSE item must default to WAREHOUSE"))
            continue
        if source_type == SupplySourceType.KITCHEN and default_source != SupplyDefaultSource.KITCHEN:
            errors.append((excel_row_no, "KITCHEN item must default to KITCHEN"))
            continue

        if item_type_text == "FINISHED":
            item_type = ItemType.finished_good
        elif item_type_text == "RAW":
            item_type = ItemType.raw_material
        else:
            errors.append((excel_row_no, f"Unsupported item type '{item_type_text}'"))
            continue

        rows.append(
            {
                "excel_row_no": excel_row_no,
                "brand_label": brand_label,
                "brand_targets": [brand_label],
                "category_name": "Final Classified Supply",
                "item_name": item_name,
                "source_type": source_type,
                "default_source": default_source,
                "kitchen_section_name": kitchen_section or None,
                "can_branch_request": source_type != SupplySourceType.NOT_REQUESTABLE,
                "visible_in_branch_ui": source_type != SupplySourceType.NOT_REQUESTABLE,
                "item_type": item_type,
                "item_code": _final_item_code(brand_label, item_name),
            }
        )

    return rows, errors


def _get_or_create_brand(db: Session, name: str) -> Brand:
    row = db.query(Brand).filter(Brand.name == name).first()
    if row:
        row.active = True
        return row
    row = Brand(name=name, active=True)
    db.add(row)
    db.flush()
    return row


def _get_or_create_category(db: Session, category_name: str) -> ItemCategory:
    code = _category_code(category_name)
    row = db.query(ItemCategory).filter(ItemCategory.code == code).first()
    if row:
        row.active = True
        row.name_ar = category_name[:100]
        row.name_en = category_name[:100]
        return row
    row = ItemCategory(
        code=code,
        name_ar=category_name[:100],
        name_en=category_name[:100],
        active=True,
    )
    db.add(row)
    db.flush()
    return row


def _get_or_create_unit(db: Session) -> UnitOfMeasure:
    row = db.query(UnitOfMeasure).filter(UnitOfMeasure.code == "PCS").first()
    if row:
        row.active = True
        return row
    row = UnitOfMeasure(code="PCS", name_ar="قطعة", name_en="Piece", active=True)
    db.add(row)
    db.flush()
    return row


def _lookup_kitchen_section(db: Session, section_name: str | None) -> KitchenSection | None:
    if not section_name:
        return None
    return db.query(KitchenSection).filter(KitchenSection.name == section_name).first()


def _item_type_for(source_type: SupplySourceType, can_branch_request: bool) -> ItemType:
    if source_type == SupplySourceType.KITCHEN:
        return ItemType.finished_good
    if can_branch_request:
        return ItemType.finished_good
    return ItemType.consumable


def _hide_unlisted_official_brand_items(
    db: Session,
    imported_item_ids: set[int],
) -> int:
    hidden = 0
    rows = (
        db.query(Item)
        .join(ItemBrand, ItemBrand.item_id == Item.id)
        .join(Brand, Brand.id == ItemBrand.brand_id)
        .filter(
            Brand.name.in_(OFFICIAL_BRANDS),
            Item.is_deleted == False,
            Item.id.notin_(imported_item_ids) if imported_item_ids else True,
        )
        .all()
    )
    for item in rows:
        changed = False
        if item.branch_requestable:
            item.branch_requestable = False
            changed = True
        if item.visible_in_branch_ui:
            item.visible_in_branch_ui = False
            changed = True
        if changed:
            hidden += 1
    return hidden


def _write_invalid_log(rejected_rows: list[dict[str, Any]], output_dir: Path) -> str:
    output_dir.mkdir(parents=True, exist_ok=True)
    csv_path = output_dir / "item_master_rejected_rows.csv"
    json_path = output_dir / "item_master_rejected_rows.json"
    fieldnames = ["excel_row_no", "reason", "item_name", "brand_label"]
    with csv_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rejected_rows:
            writer.writerow(row)
    json_path.write_text(json.dumps(rejected_rows, ensure_ascii=False, indent=2), encoding="utf-8")
    return str(csv_path)


def import_supply_item_master(
    db: Session,
    workbook_path: str | Path,
    *,
    invalid_log_dir: str | Path | None = None,
) -> ImportResult:
    workbook = Path(workbook_path)
    rows, load_errors = _load_rows(workbook)
    rows_read = len(rows) + len(load_errors)
    rejected_rows: list[dict[str, Any]] = [
        {"excel_row_no": row_no, "reason": reason}
        for row_no, reason in load_errors
    ]

    unit = _get_or_create_unit(db)
    created_items = 0
    updated_items = 0
    hidden_items = 0
    imported_item_ids: set[int] = set()

    for row in rows:
        try:
            with db.begin_nested():
                category = _get_or_create_category(db, row["category_name"])
                kitchen_section = _lookup_kitchen_section(db, row["kitchen_section_name"])
                if row["kitchen_section_name"] and kitchen_section is None:
                    raise ValueError(f"Unknown kitchen section '{row['kitchen_section_name']}'")
                item_code = row.get("item_code") or _item_code(row["brand_label"], row["category_name"], row["item_name"])

                item = db.query(Item).filter(Item.item_code == item_code).first()
                is_new = item is None
                if item is None:
                    item = Item(
                        item_code=item_code,
                        item_name_ar=row["item_name"][:200],
                        item_name_en=row["item_name"][:200],
                        category_id=category.id,
                        unit_id=unit.id,
                    )
                    db.add(item)

                item.item_name_ar = row["item_name"][:200]
                item.item_name_en = row["item_name"][:200]
                item.category_id = category.id
                item.unit_id = unit.id
                item.item_type = row.get("item_type") or _item_type_for(row["source_type"], row["can_branch_request"])
                item.storage_type = StorageType.ambient
                item.conversion_ratio = 1
                item.branch_requestable = row["can_branch_request"]
                item.visible_in_branch_ui = row["visible_in_branch_ui"]
                item.active = True
                item.source_type = row["source_type"]
                item.default_source = row["default_source"]
                item.kitchen_section_id = kitchen_section.id if kitchen_section else None
                item.is_deleted = False
                db.flush()

                mapped_brands = []
                for brand_name in row["brand_targets"]:
                    mapped_brands.append(_get_or_create_brand(db, brand_name))

                existing = {link.brand_id: link for link in item.item_brands}
                target_ids = {brand.id for brand in mapped_brands}
                for brand in mapped_brands:
                    if brand.id not in existing:
                        db.add(ItemBrand(item_id=item.id, brand_id=brand.id))
                for brand_id, link in list(existing.items()):
                    if brand_id not in target_ids:
                        db.delete(link)

                imported_item_ids.add(item.id)

                if is_new:
                    created_items += 1
                else:
                    updated_items += 1
        except Exception as exc:  # noqa: BLE001 - importer must log bad rows and continue
            rejected_rows.append(
                {
                    "excel_row_no": row["excel_row_no"],
                    "reason": str(exc),
                    "item_name": row.get("item_name"),
                    "brand_label": row.get("brand_label"),
                }
            )
            unit = _get_or_create_unit(db)

    hidden_items = _hide_unlisted_official_brand_items(db, imported_item_ids)
    db.commit()

    invalid_path = _write_invalid_log(
        rejected_rows,
        Path(invalid_log_dir or Path(__file__).resolve().parents[4] / "outputs"),
    )
    return ImportResult(
        rows_read=rows_read,
        imported_items=created_items + updated_items,
        created_items=created_items,
        updated_items=updated_items,
        hidden_items=hidden_items,
        rejected_rows=rejected_rows,
        invalid_log_path=invalid_path,
    )
